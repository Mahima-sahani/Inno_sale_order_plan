from odoo import models, fields, api
from odoo.exceptions import UserError, Warning
import logging
import openpyxl
from openpyxl.styles import Font, Alignment
import os
import shutil
import uuid
import datetime
from io import BytesIO

_logger = logging.getLogger(__name__)

class ReportWizard(models.TransientModel):

    _name = 'inno.sale.reports'
    _description = 'Sale Reports'

    def get_buyer_domain(self):
        domain = [
            ('id', 'in',
             self.env['sale.order'].search([('state', '!=', 'done'), ('state', '!=', 'cancel')]).partner_id.ids)]
        return domain

    def get_vendor_domain(self):
        vendors = self.env['dyeing.order'].search([]).mapped('partner_id')
        return [('id','in',vendors.ids)]
 

    report_type = fields.Selection(selection=[('sale_order_customize','Sale Order Customize Report'),
                                              ('sale_order_inventory_status', 'Sale Order Inventory Status'), ('report_sale_order_summary', 'Inno Sale Order Summary'),  ('report_dyeing_order_status', 'Report Dyeing Order Status')])
    to_date =  fields.Date(string="To Date")
    from_date =  fields.Date(string="From Date")
    buyer_id = fields.Many2one(comodel_name="res.partner", string="Buyer", domain=get_buyer_domain)
    division_id = fields.Many2many(comodel_name="mrp.division", string="Division")
    order_type = fields.Selection(selection=[('sale', 'Sale Order'), ('custom', 'Custom Order'),('hospitality', 'Hospitality Custom'), ('local', 'Local')], string="Order Type")
    product_group = fields.Many2many(comodel_name="product.template", string="Product Group")
    product = fields.Many2many(comodel_name="product.product", string="Product")
    planning_ids = fields.Many2many(comodel_name="inno.sale.order.planning", string="PO No.")
    excel_report = fields.Boolean(string="Excel Report")
    report_for = fields.Selection(selection=[('qty_wise','Quantity Wise'),('area_wise', 'Area Wise'),('amt_wise',"Amount Wise")], default='qty_wise', string="Report Unit")
    report_pending_all = fields.Selection(selection=[('all','All'),('pending', 'Pending')], default='all', string="Report For")
    vendor = fields.Many2one( comodel_name="res.partner",string="Vendor",domain=lambda self: self.get_vendor_domain())

    def generate_report(self):
        report = False

        if self.report_type == 'sale_order_customize':
            domain = []

            if self.buyer_id:
                domain+=[('customer_name','=',self.buyer_id.id)]
            if self.planning_ids:
                domain+=[('order_no', 'in', self.planning_ids.mapped('order_no'))]
            if self.order_type:
                domain+=[('order_type','=',self.order_type)]
            if self.from_date and self.to_date:
                domain += [('order_date', '>=', self.from_date), ('order_date', '<=', self.to_date)]
            elif self.from_date:
                domain.append(('order_date', '=', self.from_date))

            records = self.env['inno.sale.order.planning'].search(domain)

            if self.product:
                records = records.filtered(lambda rec: any(line.product_id.id in self.product.ids for line in rec.sale_order_planning_lines))

            if self.product_group:
                records = records.filtered(lambda rec: any(line.product_id.product_tmpl_id.id in self.product_group.ids for line in rec.sale_order_planning_lines))

            if self.excel_report:
                res = self.generate_excel_report(records, self.from_date, self.to_date, self.buyer_id, self.order_type)
                if res:
                    raise UserError("Excel Report has been downloaded")
            else:
                report = self.env.ref('inno_sale_order_plan.action_reports_sale_order_customize',
                                    raise_if_not_found=False).report_action(docids=records.ids,
                                                                            data={
                                                                                'to_date': self.to_date.strftime('%d/%b/%Y') if self.to_date else False,
                                                                                'from_date': self.from_date.strftime('%d/%b/%Y') if self.from_date else False,
                                                                                'docids': records.ids,
                                                                                'buyer': self.buyer_id.name if self.buyer_id else False,
                                                                                'order_type': self.order_type if self.order_type else False,
                                                                                'excel': self.excel_report if self.excel_report else False
                                                                                })
                return report
            
        if self.report_type == 'sale_order_inventory_status':
            
            domain = []
            filter = ""

            records = self.env['inno.sale.order.planning.line'].search(domain)
            if records:
                if self.division_id:
                    records = records.filtered(lambda rec: rec.product_id.division_id.id in self.division_id.ids)
                    div = [div.name for div in self.division_id]
                    st = ' '.join(div)
                    filter += f"{st}, "

                if self.from_date and self.to_date:
                    records = records.filtered(lambda rec: hasattr(rec, 'sale_order_planning_id') and rec.sale_order_planning_id and hasattr(rec.sale_order_planning_id, 'order_date') and isinstance(rec.sale_order_planning_id.order_date, datetime.date) and rec.sale_order_planning_id.order_date >= self.from_date and rec.sale_order_planning_id.order_date <= self.to_date)

                if self.product:
                    records = records.filtered(lambda rec: rec.product_id.id in self.product.ids)
                    prod = [prod.name for prod in self.product]
                    st = ' '.join(prod)
                    filter += f"{st}, "

                if self.product_group:
                    records = records.filtered(lambda rec: rec.product_id.product_tmpl_id.id in self.product_group.ids)
                    prod = [prod.name for prod in self.product_group]
                    st = ' '.join(prod)
                    filter += f"{st}, "
                    
                if self.planning_ids:
                    order_no = self.planning_ids.mapped('order_no')
                    records = records.filtered(lambda rec: rec.sale_order_planning_id.order_no in order_no)
                    po = [po.order_no for po in self.planning_ids]
                    st = ' '.join(po)
                    filter += f"{st}, "

                if self.buyer_id:
                    records = records.filtered(lambda rec: rec.sale_order_planning_id.customer_name.id == self.buyer_id.id)
                    filter += f"{self.buyer_id.name}, "

                if self.order_type:
                    records = records.filtered(lambda rec: rec.sale_order_planning_id.order_type == self.order_type)
                    filter += f"{self.order_type.capitalize()}"

                if self.excel_report:
                    res = self.inventory_status_excel_report(records, self.from_date, self.to_date, self.buyer_id, self.order_type, self.report_for, filter)
                    if res:
                        raise UserError("Excel Report has been downloaded")

                report = self.env.ref('inno_sale_order_plan.action_reports_sale_order_inventory_status',
                                        raise_if_not_found=False).report_action(docids=records.ids,
                                                                                data={
                                                                                    'docids': records.ids,
                                                                                    'to_date': self.to_date.strftime("%d/%b/%y") if self.to_date else False,
                                                                                    'from_date': self.from_date.strftime("%d/%b/%y") if self.from_date else False,
                                                                                    'report_for': self.report_for,
                                                                                    'product_id': self.product.ids,
                                                                                    'filter': filter,
                                                                                    'report_pending_all': self.report_pending_all
                                                                                    })
                return report

        if self.report_type == 'report_sale_order_summary':
            domain = []
            if self.from_date and self.to_date:
                domain += [('order_date', '>=', self.from_date), ('order_date', '<=', self.to_date)]
            records = self.env['inno.sale.order.planning'].search(domain)
            # for rec in records:

                # print('==============',rec.customer_name.name)
                

           
            report = self.env.ref('inno_sale_order_plan.action_reports_carpet_sale_order_summary',
                                raise_if_not_found=False).report_action(docids=records.ids,
                                                                        data={
                                                                            'docids':records.ids,
                                                                            'to_date': self.to_date.strftime('%d/%b/%Y') if self.to_date else False,
                                                                            'from_date': self.from_date.strftime('%d/%b/%Y') if self.from_date else False,
                                                                          
                                                                            })
            return report
        if self.report_type == 'report_dyeing_order_status':
            domain = []
            if self.from_date and self.to_date:
                domain += [('issue_date', '>=', self.from_date), ('issue_date', '<=', self.to_date),('partner_id','=',self.vendor.id)]

            
            

            records = self.env['dyeing.order'].search(domain)

           
            report = self.env.ref('inno_sale_order_plan.action_report_dyeing_order_status',
                                raise_if_not_found=False).report_action(docids=records.ids,
                                                                        data={
                                                                            'docids':records.ids,
                                                                            'to_date': self.to_date.strftime('%d/%b/%Y') if self.to_date else False,
                                                                            'from_date': self.from_date.strftime('%d/%b/%Y') if self.from_date else False,
                                                                            'vendor': self.vendor.id if self.vendor else False,
                                                                
                                                                            })
            return report

    def generate_excel_report(self, records, from_date, to_date, buyer, order_type):
        today = datetime.date.today()
        total_amount_sum = 0
        total_order_qty_sum = 0
        total_overdue_days = 0

        wb = openpyxl.Workbook()
        sheet = wb.active

        # Merge header cells and set heading
        sheet.merge_cells('A1:P2')
        heading_cell = sheet['A1']
        heading_cell.value = 'Sale Order Status Customize Summary'
        heading_cell.font = Font(size=14, bold=True)
        heading_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Insert date range text
        sheet.merge_cells('A3:P3')
        date_range_cell = sheet['A3']
        date_range_cell.value = f'Summary Type: Order No Wise Summary || From Date: {from_date} || To Date: {to_date} || Buyer: {buyer.name if buyer else False} || Document Type:  {order_type if order_type else False} || Report For:  Pending '
        date_range_cell.font = Font(bold=True)

        # Set headers
        headers = [
            "Sr No",
            "Order No",
            "Buyer Code",
            "Order Date",
            "Ship Date",
            "Over Due Days",
            "Product Name",
            "Order Qty",
            "Can Qty",
            "Order Amount",
            "Balance Qty",
            "Balance Amount",
            "Last Invoice No",
            "Last Invoice Date",
            "Last Invoice Qty",
            "Revised EDT"
        ]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=5, column=col).value = header

        # Populate data within the loop
        for index, order_line in enumerate(records.sale_order_planning_lines, start=6):
            
            due_date = order_line.sale_order_planning_id.due_date
            if due_date:
                if today < due_date:
                    days_overdue = (today - due_date).days
                else:
                    days_overdue = (today - due_date).days

            total_amount_sum += order_line.total_amount  # Add total amount to sum
            total_order_qty_sum += order_line.product_uom_qty  # Add order quantity to sum
            total_overdue_days += days_overdue if days_overdue else 0  # Add overdue days to sum

            sheet.cell(row=index, column=1).value = index - 5
            sheet.cell(row=index, column=2).value = order_line.sale_order_planning_id.order_no
            sheet.cell(row=index, column=3).value = order_line.sale_order_planning_id.customer_name.job_worker_code if order_line.sale_order_planning_id.customer_name.job_worker_code else ""
            sheet.cell(row=index, column=4).value = order_line.sale_order_planning_id.order_date.strftime('%d/%m/%y') if order_line.sale_order_planning_id.order_date else ""
            sheet.cell(row=index, column=5).value = order_line.sale_order_planning_id.due_date.strftime('%d/%m/%y') if order_line.sale_order_planning_id.due_date else ""
            sheet.cell(row=index, column=6).value = days_overdue if days_overdue else False
            sheet.cell(row=index, column=7).value = order_line.product_id.default_code
            sheet.cell(row=index, column=8).value = round(order_line.product_uom_qty, 3)
            sheet.cell(row=index, column=9).value = ""
            sheet.cell(row=index, column=10).value = round(order_line.total_amount,3)
            sheet.cell(row=index, column=11).value = ""
            sheet.cell(row=index, column=12).value = ""
            sheet.cell(row=index, column=13).value = ""
            sheet.cell(row=index, column=14).value = ""
            sheet.cell(row=index, column=15).value = ""
            sheet.cell(row=index, column=16).value = ""

        # Add total rows
        total_row_index = len(records.sale_order_planning_lines) + 7
        sheet.cell(row=total_row_index, column=1).value = "Total"
        sheet.cell(row=total_row_index, column=8).value = total_order_qty_sum
        sheet.cell(row=total_row_index, column=10).value = total_amount_sum
        sheet.cell(row=total_row_index, column=6).value = total_overdue_days

        # Bold total row
        for col in range(1, 17):  # Assuming you have 16 columns
            sheet.cell(row=total_row_index, column=col).font = Font(bold=True)

        # Set column widths
        column_widths = [6, 10, 12, 12, 12, 15, 18, 10, 10, 16, 12, 16, 14, 18, 14, 12]
        for i, width in enumerate(column_widths, start=1):
            column_letter = openpyxl.utils.get_column_letter(i)
            sheet.column_dimensions[column_letter].width = width

        # Save the Excel file in memory
        in_memory_file = BytesIO()
        wb.save(in_memory_file)
        in_memory_file.seek(0)  # Move to the beginning of the BytesIO stream

        # Generate a unique filename for download
        unique_id = uuid.uuid4()
        file_name = f"sale_order_customize_report_{unique_id}.xlsx"

        # Save the Excel file to local download folder
        download_folder = os.path.join(os.path.expanduser('~'), 'Downloads')
        file_path = os.path.join(download_folder, file_name)
        
        with open(file_path, 'wb') as f:
            f.write(in_memory_file.read())

        return file_path
    
    
    def inventory_status_excel_report(self, records, from_date, to_date, buyer, order_type, report_for, filter):
        today = datetime.date.today()

        total_sale_order_qty = 0
        total_pending_sale_order_qty = 0
        total_stock_qty = 0
        total_on_loom_qty = 0
        total_to_be_issue_qty = 0
        total_packing = 0
        if report_for == "qty_wise":
            report_unit = 'Quantity'
        elif report_for == "area_wise":
            report_unit = 'Area'
        elif report_for == "amt_wise":
            report_unit = 'Amount'

        wb = openpyxl.Workbook()
        sheet = wb.active

        # Merge header cells and set heading
        sheet.merge_cells('A1:P2')
        heading_cell = sheet['A1']
        heading_cell.value = 'Sale Order Inventory Status'
        heading_cell.font = Font(size=14, bold=True)
        heading_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Insert date range text
        sheet.merge_cells('A3:P3')
        date_range_cell = sheet['A3']
        date_range_cell.value = f'From Date: {from_date} || To Date: {to_date} || Buyer: {buyer.name if buyer else False} || Document Type:  {order_type if order_type else False} || Report For:  Pending || Report Unit: {report_unit}'
        date_range_cell.font = Font(bold=True)

        # Set headers
        headers = [
            "Sale Order No",
            "Order Date",
            "Delivery Date",
            "Over Due Days",
            "Product Name",
            "Sale Order Qty",
            "Dispatch Qty",
            "Pending Sale Order Qty",
            "Stock",
            "On Loom",
            "To Be Issue",
            "Packed Pending To Dispatch Qty",
        ]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=5, column=col).value = header

        # Populate data within the loop
        work_orders = self.env['mrp.workorder'].search([('name','=','Weaving')])
        for index, order_line in enumerate(records, start=6):

            area = order_line.product_id.mrp_area
            order_no = order_line.sale_order_planning_id.order_no
            sale_order_no = order_line.sale_order_planning_id.sale_order_id.id
            rate = order_line.rate
            
            order_no = order_line.sale_order_planning_id.order_no
            sale_order_no = order_line.sale_order_planning_id.sale_order_id.id

            packaging = self.env['stock.quant'].search([('inno_sale_id', '=', sale_order_no),('product_id','=',order_line.product_id.id),('inno_package_id','!=',False)])
            pack = len(packaging) if packaging else 0
            total_packing+= pack
            
            due_date = order_line.sale_order_planning_id.due_date
            if due_date:
                if today < due_date:
                    days_overdue = (today - due_date).days
                else:
                    days_overdue = (today - due_date).days

            work_orders = self.env['mrp.workorder'].search([('name', '=', 'Weaving'),('sale_id', '=', sale_order_no),('product_id','=',order_line.product_id.id)])
            for work_order in work_orders:
                
                if report_for == "qty_wise":
                    qty = 1
                elif report_for == "area_wise":
                    qty = area
                elif report_for == "amt_wise":
                    qty = rate

                if work_order:
                    total_sale_order_qty += order_line.product_uom_qty * qty
                    total_pending_sale_order_qty += order_line.product_uom_qty * qty
                    total_stock_qty += work_order.finished_qty * qty
                    total_to_be_issue_qty += work_order.remaining_to_allocate * qty

                    sheet.cell(row=index, column=1).value = order_no
                    sheet.cell(row=index, column=2).value = order_line.sale_order_planning_id.order_date.strftime('%d/%b/%y') if order_line.sale_order_planning_id.order_date else False
                    sheet.cell(row=index, column=3).value = order_line.sale_order_planning_id.due_date.strftime('%d/%b/%y') if order_line.sale_order_planning_id.due_date else False
                    sheet.cell(row=index, column=4).value = days_overdue if days_overdue else False
                    sheet.cell(row=index, column=5).value = order_line.product_id.default_code
                    sheet.cell(row=index, column=6).value = round(order_line.product_uom_qty * qty, 3)
                    sheet.cell(row=index, column=7).value = ""
                    sheet.cell(row=index, column=8).value = round(order_line.product_uom_qty * qty, 3)
                    sheet.cell(row=index, column=9).value = round(work_order.finished_qty * qty, 3)
                    barcode = self.env['mrp.barcode'].search([('current_process','=',work_order.id),('state','not in',['1_draft'])])
                    sheet.cell(row=index, column=10).value = round(len(barcode) * qty,3)
                    total_on_loom_qty+=(len(barcode) * qty)

                    sheet.cell(row=index, column=11).value = round(work_order.remaining_to_allocate * qty,3)
                    sheet.cell(row=index, column=12).value = pack

        # Add total rows
        total_row_index = len(records) + 7
        sheet.cell(row=total_row_index, column=1).value = "Total"
        sheet.cell(row=total_row_index, column=6).value = round(total_sale_order_qty,3)
        sheet.cell(row=total_row_index, column=8).value = round(total_pending_sale_order_qty,3)
        sheet.cell(row=total_row_index, column=9).value = round(total_stock_qty,3)

        sheet.cell(row=total_row_index, column=10).value = round(total_on_loom_qty,3)
        sheet.cell(row=total_row_index, column=11).value = round(total_to_be_issue_qty,3)
        sheet.cell(row=total_row_index, column=12).value = round(total_packing,3)

        # Bold total row
        for col in range(1, 17):  # Assuming you have 16 columns
            sheet.cell(row=total_row_index, column=col).font = Font(bold=True)

        # Set column widths
        column_widths = [14, 13, 13, 14, 15, 13, 14, 22, 8, 8, 10, 12, 30]
        for i, width in enumerate(column_widths, start=1):
            column_letter = openpyxl.utils.get_column_letter(i)
            sheet.column_dimensions[column_letter].width = width

        # Save the Excel file in memory
        in_memory_file = BytesIO()
        wb.save(in_memory_file)
        in_memory_file.seek(0)  # Move to the beginning of the BytesIO stream

        # Generate a unique filename for download
        unique_id = uuid.uuid4()
        file_name = f"sale_inventory_status_report_{unique_id}.xlsx"

        # Save the Excel file to local download folder
        download_folder = os.path.join(os.path.expanduser('~'), 'Downloads')
        file_path = os.path.join(download_folder, file_name)
        
        with open(file_path, 'wb') as f:
            f.write(in_memory_file.read())

        return file_path
    

    
